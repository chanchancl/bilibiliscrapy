
import time
import json
import openpyxl
from collections import OrderedDict


# 一个生成器，用于获得json中的所有记录
def getjson():
    with open('items.json', 'r') as f:
        while True:
            line  = f.readline()
            if line:
                yield json.loads(line, object_pairs_hook=OrderedDict)
            else:
                break

# 按照下面的顺序(列) 生成表格
Order = OrderedDict()
#  key  ,  header
#  Item的Filed Name, 在 Header显示的名字

Order['title'] = { 'Header' : '番剧名称', 'Type' : 'text'}
Order['total_view_number']      = {'Header': '总播放量', 'Type' : 'number' }
Order['total_favorite_number']  = {'Header': '追番人数', 'Type' : 'number' }
Order['total_danmaku_number']   = {'Header': '弹幕总数', 'Type' : 'number' }
Order['pub_time_text']          = {'Header': '放送时间', 'Type' : 'text' }
Order['coins']                  = {'Header': '硬币数', 'Type' : 'number' }
Order['total_episodes']         = {'Header': '总集数', 'Type' : 'number' }
Order['average_view']           = {'Header': '平均播放量', 'Type' : 'number' }
Order['average_danmaku']        = {'Header': '平均弹幕量', 'Type' : 'number' }
Order['average_coins']          = {'Header': '平均硬币数', 'Type' : 'number' }
Order['coinbiplay']             = {'Header': '币播比' , 'Type' : 'float' }
Order['total_view_text']        = {'Header': '总播放量', 'Type' : 'text' }
Order['total_favorite_text']    = {'Header': '追番人数', 'Type' : 'text' }
Order['total_danmaku_text']     = {'Header': '弹幕总数', 'Type' : 'text' }
Order['season_id']              = {'Header': 'Season ID', 'Type' : 'number' }
Order['jp_title']               = {'Header': '日文名称', 'Type' : 'text' }
Order['pub_time']               = {'Header': '放送时间(Epoch)', 'Type' : 'number' }
Order['cover_url']              = {'Header': '封面地址', 'Type' : 'text ' }

# 按播放数，降序排序
# b = sorted(data, key=lambda d : d['total_play_number'], reverse=True)

jsonData = getjson()

excFile = openpyxl.Workbook()
sheet = excFile.active

# 写入第一行
for x, key in enumerate(Order.keys()):
    sheet.cell(1, x + 1).value = Order[key]['Header']

# 遍历所有每条数据
for line ,anime in enumerate(jsonData):
    # 按Order的顺序遍历数据中的每个字段
    for column ,itemkey in enumerate(Order.keys()):
        # 获得实际值
        item = anime[itemkey]
        # 在单元格 line+1行，column列 写入数据  (空过首行)
        if itemkey.startswith('average'):
            item = round(item, 0)
        cell = sheet.cell(line+1+1, column+1)
        sheet.cell(line+1+1, column + 1).value = str(item)
        
        

filename = time.strftime("Items %Y-%m-%d(%H-%M-%S).xlsx")
excFile.save(filename)
